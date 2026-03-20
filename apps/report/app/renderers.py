from __future__ import annotations

import io
from pathlib import Path
from typing import Any, Dict

import httpx
from docxtpl import DocxTemplate
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .config import settings


class TemplateNotFoundError(FileNotFoundError):
    pass


class UnsupportedTemplateError(ValueError):
    pass


class PdfConversionError(RuntimeError):
    pass


def get_template_path(template_name: str) -> Path:
    template_path = Path(settings.report_templates_dir) / template_name
    if not template_path.exists() or not template_path.is_file():
        raise TemplateNotFoundError(f'Template not found: {template_name}')
    return template_path


def render_docx(template_name: str, data: Dict[str, Any], target_path: Path) -> None:
    template_path = get_template_path(template_name)
    if template_path.suffix.lower() != '.docx':
        raise UnsupportedTemplateError('DOCX rendering requires a .docx template')

    doc = DocxTemplate(str(template_path))
    doc.render(data)
    doc.save(str(target_path))


async def convert_docx_to_pdf(source_path: Path, target_path: Path) -> None:
    async with httpx.AsyncClient(timeout=120.0) as client:
        with source_path.open('rb') as file_handle:
            files = {
                'files': (
                    source_path.name,
                    file_handle,
                    'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                )
            }
            response = await client.post(
                f'{settings.gotenberg_url}/forms/libreoffice/convert',
                files=files,
            )

    if response.status_code != 200:
        raise PdfConversionError(f'Gotenberg error: {response.status_code} {response.text}')

    target_path.write_bytes(response.content)


def render_xlsx(template_name: str, data: Dict[str, Any], target_path: Path) -> None:
    template_path = get_template_path(template_name)
    if template_path.suffix.lower() != '.json':
        raise UnsupportedTemplateError('XLSX rendering requires a .json template descriptor')

    import json

    descriptor = json.loads(template_path.read_text(encoding='utf-8'))
    wb = Workbook()
    ws = wb.active
    ws.title = descriptor.get('sheet_title', 'Report')

    title = descriptor.get('title', 'Report')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='left')

    row = 3
    for item in descriptor.get('fields', []):
        label = item['label']
        key = item['key']
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=_resolve_value(data, key))
        row += 1

    for table in descriptor.get('tables', []):
        row += 1
        ws.cell(row=row, column=1, value=table['title']).font = Font(bold=True, size=12)
        row += 1
        headers = table['columns']
        for col_idx, column in enumerate(headers, start=1):
            ws.cell(row=row, column=col_idx, value=column['label']).font = Font(bold=True)
        row += 1
        records = _resolve_value(data, table['key']) or []
        for record in records:
            for col_idx, column in enumerate(headers, start=1):
                ws.cell(row=row, column=col_idx, value=_resolve_value(record, column['key']))
            row += 1

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 42
    wb.save(str(target_path))


def _resolve_value(data: Dict[str, Any], dotted_key: str) -> Any:
    current: Any = data
    for part in dotted_key.split('.'):
        if isinstance(current, dict):
            current = current.get(part)
        else:
            return None
    return current
